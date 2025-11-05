#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 파일을 JSON으로 변환하는 스크립트
큐플족보 Excel 파일의 질문/답 데이터를 JSON 형식으로 변환합니다.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("오류: openpyxl이 설치되지 않았습니다.")
    print("다음 명령어로 설치해주세요: pip install openpyxl")
    sys.exit(1)


def parse_excel_sheet(worksheet, sheet_name):
    """
    Excel 워크시트를 파싱하여 질문/답 데이터 추출
    
    Args:
        worksheet: openpyxl 워크시트 객체
        sheet_name: 시트 이름
    
    Returns:
        list: 질문/답 데이터 리스트
    """
    data = []
    
    # 워크시트의 모든 행 가져오기
    rows = list(worksheet.iter_rows(values_only=True))
    
    if not rows or len(rows) == 0:
        return data
    
    # 첫 번째 행이 헤더인지 확인
    first_row = rows[0]
    has_header = False
    question_col = 0
    answer_col = 1
    
    # 자동으로 문제/답 컬럼 찾기
    if first_row and len(first_row) > 0:
        headers = [str(h or '').lower() for h in first_row]
        
        # 헤더에서 문제/답 컬럼 찾기
        question_keywords = ['문제', 'question', '질문', '문항', '문제내용', '문항내용', '내용']
        answer_keywords = ['답', 'answer', '정답', '해답', '답안', '해설', '정답내용']
        
        for idx, header in enumerate(headers):
            if any(keyword in header for keyword in question_keywords):
                question_col = idx
                has_header = True
            if any(keyword in header for keyword in answer_keywords):
                answer_col = idx
                has_header = True
        
        # 헤더가 없으면 첫 번째 열을 문제, 두 번째 열을 답으로 간주
        if not has_header:
            question_col = 0
            answer_col = 1
    
    # 데이터 행 처리
    start_row = 1 if has_header else 0
    for i in range(start_row, len(rows)):
        row = rows[i]
        if not row or len(row) == 0:
            continue
        
        # 열 인덱스가 범위를 벗어나지 않도록 확인
        question = str(row[question_col] if question_col < len(row) else '').strip()
        answer = str(row[answer_col] if answer_col < len(row) else '').strip()
        
        if question:
            data.append({
                'question': question,
                'answer': answer or '답이 없습니다.',
                'index': i - start_row + 1,
                'sheet': sheet_name or '알 수 없음'
            })
    
    return data


def convert_excel_to_json(excel_path, output_path=None):
    """
    Excel 파일을 JSON으로 변환
    
    Args:
        excel_path: Excel 파일 경로
        output_path: 출력 JSON 파일 경로 (None이면 자동 생성)
    """
    excel_path = Path(excel_path)
    
    if not excel_path.exists():
        print(f"오류: 파일을 찾을 수 없습니다: {excel_path}")
        sys.exit(1)
    
    print(f"Excel 파일 읽는 중: {excel_path}")
    
    try:
        # Excel 파일 열기
        workbook = openpyxl.load_workbook(excel_path, data_only=True)
        sheet_names = workbook.sheetnames
        
        print(f"시트 개수: {len(sheet_names)}")
        print(f"시트 목록: {', '.join(sheet_names)}")
        
        # 모든 시트 데이터 수집
        all_data = []
        for sheet_name in sheet_names:
            worksheet = workbook[sheet_name]
            sheet_data = parse_excel_sheet(worksheet, sheet_name)
            all_data.extend(sheet_data)
            print(f"  - {sheet_name}: {len(sheet_data)}개 항목")
        
        print(f"\n총 {len(all_data)}개의 문제를 찾았습니다.")
        
        # 출력 경로 설정
        if output_path is None:
            output_path = excel_path.parent / 'data.json'
        else:
            output_path = Path(output_path)
        
        # JSON 파일로 저장
        output_data = {
            'metadata': {
                'source_file': str(excel_path.name),
                'total_items': len(all_data),
                'sheets': sheet_names,
                'sheet_count': len(sheet_names)
            },
            'data': all_data
        }
        
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(output_data, f, ensure_ascii=False, indent=2)
        
        print(f"\nJSON 파일 생성 완료: {output_path}")
        print(f"파일 크기: {output_path.stat().st_size / 1024:.2f} KB")
        
    except Exception as e:
        print(f"오류 발생: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


def main():
    """메인 함수"""
    # 기본 Excel 파일 경로
    default_excel = Path('큐플족보_251021.xlsx')
    
    if len(sys.argv) > 1:
        excel_path = sys.argv[1]
    else:
        excel_path = default_excel
    
    if len(sys.argv) > 2:
        output_path = sys.argv[2]
    else:
        output_path = None
    
    convert_excel_to_json(excel_path, output_path)


if __name__ == '__main__':
    main()




